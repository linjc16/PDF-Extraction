import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from tqdm import tqdm
import os

import pdb


def visualize_to_excel(save_path, fig_path_list, catpion_list, html_list, passage_list):

    # Open workbook and select worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # 将标题行写入工作表
    header = ['Image', 'caption', 'passage', 'html']
    for col_num, column_title in enumerate(header, 1):
        column_letter = get_column_letter(col_num)
        worksheet[f'{column_letter}1'] = column_title

    image_max_height = 0
    

    for i in tqdm(range(min(300, len(fig_path_list)))):
        try:
            # Load image
            pil_image = PILImage.open(fig_path_list[i])
            
            # 计算缩放比例
            max_width = 400  # 假设最大宽度是300个像素
            max_height = pil_image.height  # 假设最大高度是200个像素
            width_ratio = max_width / pil_image.width
            height_ratio = max_height / pil_image.height
            ratio = min(width_ratio, height_ratio)

            # 缩放图像
            new_width = int(pil_image.width * ratio)
            new_height = int(pil_image.height * ratio)
            pil_image = pil_image.resize((new_width, new_height), PILImage.LANCZOS)

            # Add image to worksheet
            xl_image = XLImage(fig_path_list[i])
            xl_image.width = new_width
            xl_image.height = new_height
            xl_image.anchor = f'A{str(i + 2)}'  # 设置图像元素的锚定
            worksheet.add_image(xl_image, f'A{str(i + 2)}')

            # 写入标题和段落
            worksheet[f'B{str(i + 2)}'] = catpion_list[i]
            worksheet[f'C{str(i + 2)}'] = passage_list[i]
            worksheet[f'D{str(i + 2)}'] = html_list[i]

            length_max = 0
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                length_max = max(length_max, length)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = 40
            
            # Get dimensions of the image
            img_width, img_height = pil_image.size
            
            # Set column width to fit image width
            col_width = img_width * 0.145
            col_letter = get_column_letter(1)
            worksheet.column_dimensions[col_letter].width = col_width

            image_max_height = max(image_max_height, new_height)

            worksheet.row_dimensions[i + 2].height = max(700, new_height)
        
        except:
            continue

    for row in worksheet:
        for cell in row:
            if cell.value is not None:
                cell.alignment = cell.alignment.copy(wrap_text=True, horizontal='left', vertical='center')

    
    
    # Save workbook
    workbook.save(os.path.join(save_path, 'output.xlsx'))